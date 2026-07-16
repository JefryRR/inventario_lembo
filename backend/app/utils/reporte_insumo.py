import io
from openpyxl import Workbook   # type: ignore
from openpyxl.styles import Font, PatternFill  # type: ignore
from reportlab.lib import colors  # type: ignore
from reportlab.lib.pagesizes import letter, landscape  # type: ignore
from reportlab.lib.units import cm # type: ignore
from reportlab.lib.styles import getSampleStyleSheet  # type: ignore
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer # type: ignore
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.styles import ParagraphStyle

styles = getSampleStyleSheet()

estilo_observaciones = styles["BodyText"].clone("Observaciones")
estilo_observaciones.fontSize = 6
estilo_observaciones.leading = 7


def generar_excel_reporte_insumo(reporte: dict) -> io.BytesIO:
    encabezado = reporte["encabezado"]
    movimientos = reporte["movimientos"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Encabezado"

    ws.append(["Informe de Insumo", encabezado["nombre_producto"]])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])
    ws.append(["ID insumo", encabezado["id_insumo"]])
    ws.append(["Fecha de ingreso", str(encabezado["fecha_ingreso"])])
    ws.append(["Fecha de vencimiento", str(encabezado["fecha_vencimiento"])])
    ws.append(["Precio unitario", float(encabezado["precio_unitario"])])
    ws.append(["Cantidad inicial", encabezado["cantidad_inicial"]])
    ws.append(["Stock actual", encabezado["stock_actual"]])
    ws.append(["Cantidad solicitada", encabezado["total_solicitado"]])
    ws.append(["Cantidad devuelta", encabezado["total_devuelto"]])
    ws.append(["Total perdido", encabezado["total_insumo"]])
    ws.append(["Valor total del insumo", f"${encabezado['cantidad_inicial']* encabezado['precio_unitario']:,.0f}"])

    for fila in ws.iter_rows(min_row=3, max_row=12, min_col=1, max_col=1):
        for celda in fila:
            celda.font = Font(bold=True)

    ws_mov = wb.create_sheet("Movimientos")
    ws_mov.append(["Tipo", "Observaciones", "Cantidad", "Unidad", "Valor", "Motivo", "Fecha"])
    for celda in ws_mov[1]:
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = PatternFill(start_color="007832", end_color="007832", fill_type="solid")

    for m in movimientos:
        ws_mov.append([
            m.get("tipo"),
            m.get("observaciones"),
            m.get("cantidad"),
            encabezado.get("simbolo"),
            float(m["valor"]) if m.get("valor") not in (None, "-") else "-",
            m.get("motivo"),
            str(m.get("fecha")),
        ])

    for columna in ws_mov.columns:
        max_len = max((len(str(c.value)) if c.value else 0) for c in columna)
        ws_mov.column_dimensions[columna[0].column_letter].width = max_len + 2

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

def generar_pdf_reporte_insumo(reporte: dict) -> io.BytesIO:
    encabezado = reporte["encabezado"]
    movimientos = reporte["movimientos"]

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    styles = getSampleStyleSheet()

    # Estilo centrado para el título de historial de estados
    estilo_titulo_centrado = ParagraphStyle(
        "TituloCentrado",
        parent=styles["Heading2"],
        alignment=TA_CENTER,
    )
    elementos = []

    elementos.append(Paragraph(f"Informe de Insumo: {encabezado['nombre_producto']}", styles["Title"]))
    elementos.append(Spacer(1, 12))

    datos_encabezado = [
        ["ID insumo", str(encabezado["id_insumo"])],
        ["Nombre del insumo", str(encabezado["nombre_producto"])],
        ["Fecha de ingreso", str(encabezado["fecha_ingreso"])],
        ["Fecha de vencimiento", str(encabezado["fecha_vencimiento"])],
        ["Precio unitario", f"${float(encabezado['precio_unitario']):,.0f}"],
        ["Cantidad inicial", f"{encabezado['cantidad_inicial']} {encabezado['simbolo']}"],
        ["Stock actual", f"{encabezado['stock_actual']} {encabezado['simbolo']}"],
        ["Total perdido", f"{encabezado['total_insumo']} {encabezado['simbolo']}"],
        ["Total solicitado", f"{encabezado['total_solicitado']} {encabezado['simbolo']}"],
        ["Total devuelto", f"{encabezado['total_devuelto']} {encabezado['simbolo']}"],
        ["Valor total del insumo", f"${encabezado['cantidad_inicial']* encabezado['precio_unitario']:,.0f}"],

    ]
    tabla_encabezado = Table(datos_encabezado, colWidths=[6 * cm, 6 * cm])
    tabla_encabezado.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, -1), colors.whitesmoke),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
    ]))
    elementos.append(tabla_encabezado)
    elementos.append(Spacer(1, 20))

    elementos.append(Paragraph("Movimientos", estilo_titulo_centrado))
    filas = [["Tipo", "Observaciones", "Cantidad", "Valor", "Motivo", "Fecha"]]
    for m in movimientos:
        valor = m.get("valor")
        filas.append([
            m.get("tipo", ""),
            Paragraph(m.get("observaciones") or "", estilo_observaciones),
            str(m.get("cantidad", "")),
            f"${float(valor):,.0f}" if valor not in (None, "-") else "-",
            m.get("motivo") or "",
            str(m.get("fecha", "")),
        ])

    tabla_movs = Table(
        filas,
        repeatRows=1,
        colWidths=[2.3 * cm, 6 * cm, 2 * cm, 2.3 * cm, 2.7 * cm, 2.3 * cm],
    )
    tabla_movs.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f3f4f6")),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))
    elementos.append(tabla_movs)

    doc.build(elementos)
    buffer.seek(0)
    return buffer

def generar_pdf_rep_gral_insumos(reporte: list) -> io.BytesIO:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(letter),
        leftMargin=1.5 * cm,
        rightMargin=1.5 * cm,
        topMargin=1.5 * cm,
        bottomMargin=1.5 * cm,
    )
    styles = getSampleStyleSheet()
    elementos = []

    elementos.append(Paragraph("Informe de insumos", styles["Title"]))
    elementos.append(Spacer(1, 12))

    total_insumo = 0.0
    filas = [["Producto", "cantidad", "Tipo insumo", "F. registro", "F. vencimiento", "Estado stock", "Precio unit.", "Precio total.", "Detalles"]]
    for i in reporte:
        valor_unitario = i.get("precio_unitario")
        valor_total = _calcular_valor_total_ins(i)
        total_insumo += valor_total
        estado_stock = "Agotado" if i.get("cantidad") < 20 else "Provisionar" if i.get("cantidad") < i.get("min_stock") else "Disponible"
        filas.append([
            Paragraph(i.get("nombre_producto") or "", estilo_observaciones),
            i.get("cantidad") or "",
            Paragraph(i.get("nombre_tipo") or "", estilo_observaciones),
            i.get("fecha_ingreso") or "",
            i.get("fecha_vencimiento") or "",
            Paragraph(estado_stock, estilo_observaciones),
            f"${float(valor_unitario):,.0f}" if valor_unitario not in (None, "-") else "-",
            f"${valor_total:,.0f}",
            Paragraph(i.get("nivel_alerta") or "Sin detalles", estilo_observaciones)
        ])

    tabla = Table(
        filas,
        repeatRows=1,
        colWidths=[
            2.5 * cm,
            2.3 * cm,
            2 * cm,
            2.7 * cm,
            2.7 * cm,
            2 * cm,
            2 * cm,
            2.5 * cm,
            2 * cm,
            4 * cm,
        ]
    )
    tabla.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f3f4f6")),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]))
    elementos.append(tabla)
    elementos.append(Spacer(1, 18))

    resumen = [
        ["Total de registros", str(len(reporte))],
        ["Valor total de insumos", f"${total_insumo:,.0f}"],
    ]
    tabla_resumen = Table(resumen, colWidths=[6 * cm, 6 * cm])
    tabla_resumen.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, -1), colors.whitesmoke),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
    ]))
    elementos.append(tabla_resumen)

    doc.build(elementos)
    buffer.seek(0)
    return buffer

def generar_excel_rep_gral_insumos(insumos: list) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Insumos"

    ws.append(["Informe de Insumos"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([f"Total de registros: {len(insumos)}"])
    ws.append([])

    headers = [
        "Producto", "Cantidad", "Tipo insumo", "Fecha Registro", "Fecha vencimiento",  
        "Estado stock", "Precio unitario", "Precio total"
    ]
    ws.append(headers)
    fila_encabezado = ws.max_row
    for celda in ws[fila_encabezado]:
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = PatternFill(start_color="007832", end_color="007832", fill_type="solid")
    total_insumo = 0.0
    for i in insumos:
        valor_total = _calcular_valor_total_ins(i)
        total_insumo += valor_total
        estado_stock = "Agotado" if i.get("cantidad") < 20 else "Provisionar" if i.get("cantidad") < i.get("min_stock") else "Disponible"

        ws.append([
            i.get("nombre_producto"),
            f'{i.get("cantidad") or "0"} {i.get("simbolo") or ""}'.strip(),
            i.get("nombre_tipo") or "-",
            i.get("fecha_ingreso"),
            i.get("fecha_vencimiento"),
            estado_stock,
            float(i.get("precio_unitario")) if i.get("precio_unitario") not in (None, "-") else "-",
            valor_total
        ])

    ws.append([])
    fila_total = ws.max_row + 1
    ws.cell(row=fila_total, column=8, value="Valor total insumo:").font = Font(bold=True)
    ws.cell(row=fila_total, column=9, value=total_insumo).font = Font(bold=True)

    for columna in ws.columns:
        max_len = max((len(str(c.value)) if c.value else 0) for c in columna)
        ws.column_dimensions[columna[0].column_letter].width = max_len + 2

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

def _calcular_valor_total_ins(insumo: dict) -> float:
    valor_unitario = insumo.get("precio_unitario")
    cantidad = insumo.get("cantidad")
    if valor_unitario in (None, "-") or cantidad in (None, "-"):
        return 0.0
    return float(valor_unitario) * float(cantidad)
