import io
from openpyxl import Workbook   # type: ignore
from openpyxl.styles import Font, PatternFill  # type: ignore
from reportlab.lib import colors  # type: ignore
from reportlab.lib.pagesizes import letter, landscape  # type: ignore
from reportlab.lib.units import cm # type: ignore
from reportlab.lib.styles import getSampleStyleSheet  # type: ignore
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer # type: ignore
from collections import defaultdict
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.styles import ParagraphStyle

styles = getSampleStyleSheet()

estilo_observaciones = styles["BodyText"].clone("Observaciones")
estilo_observaciones.fontSize = 6
estilo_observaciones.leading = 7


def _normalizar_reporte_maquina(reporte: dict) -> tuple[dict, list[dict]]:
    historial = reporte.get("historial") or reporte.get("movimientos") or []

    if reporte.get("encabezado"):
        encabezado = reporte["encabezado"]
    else:
        encabezado = historial[0] if historial else {}

    return encabezado, historial

def _calcular_valor_total(perdida: dict) -> float:
    valor_unitario = perdida.get("valor_unitario")
    cantidad = perdida.get("cantidad")
    if valor_unitario in (None, "-") or cantidad in (None, "-"):
        return 0.0
    return float(valor_unitario) * float(cantidad)

def generar_excel_reporte_perdidas(perdidas: list) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Pérdidas"

    ws.append(["Informe de Pérdidas"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([f"Total de registros: {len(perdidas)}"])
    ws.append([])

    headers = [
        "Producto", "Origen", "Lote", "Motivo",
        "Cantidad", "Unidad", "Valor unitario", "Valor total perdido",
        "Usuario", "Fecha de reporte", "Observaciones",
    ]
    ws.append(headers)
    fila_encabezado = ws.max_row
    for celda in ws[fila_encabezado]:
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = PatternFill(start_color="007832", end_color="007832", fill_type="solid")

    total_perdido = 0.0
    for p in perdidas:
        valor_unitario = p.get("valor_unitario")
        valor_total = _calcular_valor_total(p)
        total_perdido += valor_total

        ws.append([
            p.get("nombre_producto"),
            p.get("origen"),
            p.get("nombre_lote") or "-",
            p.get("motivo"),
            f'{p.get("cantidad") or "0"} {p.get("simbolo") or ""}'.strip(),
            p.get("simbolo"),
            float(valor_unitario) if valor_unitario not in (None, "-") else "-",
            valor_total,
            p.get("nombre_user") or "Sistema",
            str(p.get("fecha_reporte")),
            p.get("observaciones") or "",
        ])

    ws.append([])
    fila_total = ws.max_row + 1
    ws.cell(row=fila_total, column=8, value="Total perdido:").font = Font(bold=True)
    ws.cell(row=fila_total, column=9, value=total_perdido).font = Font(bold=True)

    for columna in ws.columns:
        max_len = max((len(str(c.value)) if c.value else 0) for c in columna)
        ws.column_dimensions[columna[0].column_letter].width = max_len + 2

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

def generar_pdf_reporte_perdidas(perdidas: list) -> io.BytesIO:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(letter),
        leftMargin=1.5 * cm,
        rightMargin=1.5 * cm,
        topMargin=1.5 * cm,
        bottomMargin=1.5 * cm,
    )
    styles = getSampleStyleSheet()
    elementos = []

    elementos.append(Paragraph("Informe de Pérdidas", styles["Title"]))
    elementos.append(Spacer(1, 12))

    total_perdido = 0.0
    filas = [["Producto", "Origen", "Lote", "Motivo", "Observaciones", "Cantidad", "Valor unit.", "Valor total", "Usuario", "Fecha"]]
    for p in perdidas:
        valor_unitario = p.get("valor_unitario")
        valor_total = _calcular_valor_total(p)
        total_perdido += valor_total
        cantidad = p.get("cantidad")
        simbolo = p.get("simbolo") or ""

        filas.append([
            p.get("nombre_producto") or "",
            p.get("origen") or "",
            p.get("nombre_lote") or "-",
            p.get("motivo") or "",
            Paragraph(p.get("observaciones") or "", estilo_observaciones),
            f"{cantidad} {simbolo}".strip(),
            f"${float(valor_unitario):,.0f}" if valor_unitario not in (None, "-") else "-",
            f"${valor_total:,.0f}",
            p.get("nombre_user") or "Sistema",
            str(p.get("fecha_reporte", "")),
        ])

    tabla = Table(
        filas,
        repeatRows=1,
        colWidths=[
            2.5 * cm,
            2.3 * cm,
            2 * cm,
            2.7 * cm,
            4 * cm,
            2 * cm,
            2 * cm,
            2.5 * cm,
            2 * cm,
            3 * cm,
        ]
    )
    tabla.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f3f4f6")),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]))
    elementos.append(tabla)
    elementos.append(Spacer(1, 18))

    resumen = [
        ["Total de registros", str(len(perdidas))],
        ["Total valor perdido", f"${total_perdido:,.0f}"],
    ]
    tabla_resumen = Table(resumen, colWidths=[6 * cm, 6 * cm])
    tabla_resumen.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, -1), colors.whitesmoke),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
    ]))
    elementos.append(tabla_resumen)

    doc.build(elementos)
    buffer.seek(0)
    return buffer

def generar_excel_reporte_lote_prod(reporte: dict) -> io.BytesIO:
    encabezado = reporte["encabezado"]
    historial = reporte["historial"]
    mortalidades = reporte["mortalidades"]

    wb = Workbook()

    # ── Hoja 1: Encabezado ──────────────────────────────────────────
    ws = wb.active
    ws.title = "Resumen"

    titulo = ws["A1"]
    titulo.value = f"Informe de lote: {encabezado['nombre_lote']}"
    titulo.font = Font(bold=True, size=14)
    ws.append([])

    campos = [
        ("Especie",               encabezado["nombre_especie"]),
        ("Categoría",             encabezado["nombre_categoria"]),
        ("Fecha de siembra",      str(encabezado["fecha_siembra"])),
        ("Fecha de cosecha",      str(encabezado["fecha_cosecha"])),
        ("Estado",                encabezado["estado_lote"]),
        ("Responsable",           encabezado["nombre_user"]),
        ("Cantidad inicial",      encabezado["cantidad_inicial"]),
        ("Vivos actuales",        encabezado["cantidad"]),
        ("Total muertes",         encabezado["total_muertes"]),
        ("% Mortalidad",          f"{encabezado['porcentaje_mortalidad']}%"),
    ]
    for etiqueta, valor in campos:
        fila = ws.append([etiqueta, valor])
    # Negrita en etiquetas
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=1):
        for cell in row:
            cell.font = Font(bold=True)

    # ── Hoja 2: Historial de estados ────────────────────────────────
    ws_hist = wb.create_sheet("Historial estados")
    headers = ["Fecha de cambio", "Estado", "Registrado por"]
    ws_hist.append(headers)
    for cell in ws_hist[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="007832", end_color="007832", fill_type="solid")

    for h in historial:
        ws_hist.append([
            str(h.get("fecha_cambio", "")),
            h.get("estado", ""),
            h.get("nombre_user", ""),
        ])

    for col in ws_hist.columns:
        max_len = max((len(str(c.value)) if c.value else 0) for c in col)
        ws_hist.column_dimensions[col[0].column_letter].width = max_len + 4

    # ── Hoja 3: Mortalidades ─────────────────────────────────────────
    ws_mort = wb.create_sheet("Mortalidades")
    headers_mort = ["Fecha de reporte", "Cantidad", "Observación", "Registrado por"]
    ws_mort.append(headers_mort)
    for cell in ws_mort[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="DC2626", end_color="DC2626", fill_type="solid")

    for m in mortalidades:
        ws_mort.append([
            str(m.get("fecha_reporte", "")),
            m.get("cantidad", 0),
            m.get("observacion", "") or "",
            m.get("nombre_user", "") or "",
        ])

    for col in ws_mort.columns:
        max_len = max((len(str(c.value)) if c.value else 0) for c in col)
        ws_mort.column_dimensions[col[0].column_letter].width = max_len + 4

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

def generar_pdf_reporte_lotes_prod(reporte: dict) -> io.BytesIO:
    encabezado  = reporte["encabezado"]
    historial   = reporte["historial"]
    mortalidades = reporte["mortalidades"]

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    styles = getSampleStyleSheet()
    elementos = []

    # ── Título ───────────────────────────────────────────────────────
    elementos.append(Paragraph(
        f"Informe de lote: {encabezado['nombre_lote']}", styles["Title"]
    ))
    elementos.append(Spacer(1, 12))

    # ── Encabezado general ───────────────────────────────────────────
    datos_enc = [
        ["Especie",            str(encabezado["nombre_especie"])],
        ["Categoría",          str(encabezado["nombre_categoria"])],
        ["Fecha de siembra",   str(encabezado["fecha_siembra"])],
        ["Fecha de cosecha",   str(encabezado["fecha_cosecha"])],
        ["Estado",             encabezado["estado_lote"]],
        ["Responsable",        encabezado["nombre_user"]],
    ]
    tabla_enc = Table(datos_enc, colWidths=[6 * cm, 9 * cm])
    tabla_enc.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, -1), colors.whitesmoke),
        ("FONTNAME",   (0, 0), (0, -1), "Helvetica-Bold"),
        ("GRID",       (0, 0), (-1, -1), 0.5, colors.grey),
        ("FONTSIZE",   (0, 0), (-1, -1), 9),
    ]))
    elementos.append(tabla_enc)
    elementos.append(Spacer(1, 16))

    # ── Métricas ─────────────────────────────────────────────────────
    elementos.append(Paragraph("Resumen de mortalidad", styles["Heading2"]))
    datos_metricas = [
        ["Cantidad inicial", "Vivos actuales", "Total muertes", "% Mortalidad"],
        [
            str(encabezado["cantidad_inicial"]),
            str(encabezado["cantidad"]),
            str(encabezado["total_muertes"]),
            f"{encabezado['porcentaje_mortalidad']}%",
        ],
    ]
    tabla_metricas = Table(datos_metricas, colWidths=[4 * cm] * 4)
    tabla_metricas.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1e40af")),
        ("TEXTCOLOR",  (0, 0), (-1, 0), colors.white),
        ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ALIGN",      (0, 0), (-1, -1), "CENTER"),
        ("GRID",       (0, 0), (-1, -1), 0.5, colors.grey),
        ("FONTSIZE",   (0, 0), (-1, -1), 9),
    ]))
    elementos.append(tabla_metricas)
    elementos.append(Spacer(1, 20))

    # ── Historial de estados ─────────────────────────────────────────
    elementos.append(Paragraph("Historial de estados", styles["Heading2"]))
    filas_hist = [["Fecha de cambio", "Estado", "Registrado por"]]
    for h in historial:
        filas_hist.append([
            str(h.get("fecha_cambio", "")),
            h.get("estado", ""),
            h.get("nombre_user", "") or "-",
        ])

    tabla_hist = Table(filas_hist, repeatRows=1, colWidths=[5 * cm, 5 * cm, 5 * cm])
    tabla_hist.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f3f4f6")),
        ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",   (0, 0), (-1, -1), 8),
        ("GRID",       (0, 0), (-1, -1), 0.5, colors.grey),
        ("VALIGN",     (0, 0), (-1, -1), "MIDDLE"),
    ]))
    elementos.append(tabla_hist)
    elementos.append(Spacer(1, 20))

    # ── Mortalidades ─────────────────────────────────────────────────
    elementos.append(Paragraph("Detalle de mortalidades", styles["Heading2"]))
    filas_mort = [["Fecha de reporte", "Cantidad", "Observación", "Registrado por"]]
    for m in mortalidades:
        filas_mort.append([
            str(m.get("fecha_reporte", "")),
            str(m.get("cantidad", "")),
            m.get("observacion", "") or "-",
            m.get("nombre_user", "") or "-",
        ])

    tabla_mort = Table(filas_mort, repeatRows=1, colWidths=[4*cm, 3*cm, 6*cm, 4*cm])
    tabla_mort.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#fef2f2")),
        ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",   (0, 0), (-1, -1), 8),
        ("GRID",       (0, 0), (-1, -1), 0.5, colors.grey),
        ("VALIGN",     (0, 0), (-1, -1), "MIDDLE"),
    ]))
    elementos.append(tabla_mort)

    doc.build(elementos)
    buffer.seek(0)
    return buffer

def generar_excel_reporte_mortalidad(perdidas: list) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Mortalidad"

    ws.append(["Informe de mortalidad"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([f"Total de registros: {len(perdidas)}"])
    ws.append([])

    headers = [
        "Lote", "Sublote", "Categoría", "especie", "Cantidad", "Fecha reporte", "Observación", "usuario"
    ]
    ws.append(headers)
    fila_encabezado = ws.max_row
    for celda in ws[fila_encabezado]:
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = PatternFill(start_color="007832", end_color="007832", fill_type="solid")

    total_mortalidad = 0.0
    for p in perdidas:
       
        ws.append([ # type: ignore
            p.get("nombre_lote") if p.get("nombre_lote") else "-",
            p.get("sublote") if p.get("sublote") else "-",
            p.get("nombre_categoria") if p.get("nombre_categoria") else "-",
            p.get("nombre_especie") if p.get("nombre_especie") else "-",
            p.get("cantidad"),
            str(p.get("fecha_reporte", "")),
            p.get("observacion") or "",
            p.get("nombre_user") if p.get("nombre_user") else "-",
        ])
        total_mortalidad += p.get("cantidad", 0)
    ws.append([]) # type: ignore
    fila_total = ws.max_row + 1
    ws.cell(row=fila_total, column=4, value="Total mortalidad:").font = Font(bold=True)
    ws.cell(row=fila_total, column=5, value=total_mortalidad).font = Font(bold=True)

    for columna in ws.columns:
        max_len = max((len(str(c.value)) if c.value else 0) for c in columna)
        ws.column_dimensions[columna[0].column_letter].width = max_len + 2

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

def generar_pdf_reporte_mortalidad(mortalidad: list) -> io.BytesIO:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(letter),
        leftMargin=1.5 * cm,
        rightMargin=1.5 * cm,
        topMargin=1.5 * cm,
        bottomMargin=1.5 * cm,
    )
    styles = getSampleStyleSheet()
    elementos = []

    elementos.append(Paragraph("Informe de mortalidad", styles["Title"]))
    elementos.append(Spacer(1, 12))

    total_mortalidad = 0.0
    filas = [["Lote", "Sublote", "Categoría", "especie", "Cantidad", "Fecha reporte", "Observación", "usuario"]]
    for p in mortalidad:
        total_mortalidad += p.get("cantidad", 0)

        filas.append([
            p.get("nombre_lote") or "",
            p.get("sublote") or "",
            p.get("nombre_categoria") or "",
            p.get("nombre_especie") or "",
            p.get("cantidad", 0),
            str(p.get("fecha_reporte", "")),
            Paragraph(p.get("observacion") or "", estilo_observaciones),
            p.get("nombre_user") or "-",
        ])

    tabla = Table(
        filas,
        repeatRows=1,
        colWidths=[
            2.3 * cm,
            2 * cm,
            2 * cm,
            2 * cm,
            3 * cm,
            4 * cm,
            3 * cm,
        ]
    )
    tabla.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f3f4f6")),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]))
    elementos.append(tabla)
    elementos.append(Spacer(1, 18))

    resumen = [
        ["Total de registros", str(len(mortalidad))],
        ["Total mortalidad", f"{total_mortalidad:,.0f}"],
    ]
    tabla_resumen = Table(resumen, colWidths=[6 * cm, 6 * cm])
    tabla_resumen.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, -1), colors.whitesmoke),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
    ]))
    elementos.append(tabla_resumen)

    doc.build(elementos)
    buffer.seek(0)
    return buffer

def _agrupar_detalles_por_venta(detalles: list) -> dict:
    agrupado = defaultdict(list)
    for d in detalles:
        agrupado[d["venta_id"]].append(d)
    return agrupado

def generar_excel_reporte_ventas(ventas: list, detalles: list) -> io.BytesIO:
    detalles_por_venta = _agrupar_detalles_por_venta(detalles)

    wb = Workbook()
    ws = wb.active
    ws.title = "Historial de Ventas"

    ws.append(["Informe de Ventas"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([f"Total de ventas: {len(ventas)}"])
    ws.append([])

    total_general = 0.0

    for venta in ventas:
        total_venta = float(venta.get("total_venta") or 0)
        total_general += total_venta

        ws.append([
            f"Venta #{venta['id_venta']}",
            venta.get("nombre_comprador") or "-",
            str(venta.get("fecha_venta")),
            venta.get("nombre_user") or "-",
            f"Total: ${total_venta:,.0f}",
        ])
        fila_resumen = ws.max_row
        for celda in ws[fila_resumen]:
            celda.font = Font(bold=True, color="FFFFFF")
            celda.fill = PatternFill(start_color="00304D", end_color="00304D", fill_type="solid")

        ws.append(["Producto", "Cantidad", "Unidad", "Precio unitario", "Precio total", "Estado"])
        fila_encabezado = ws.max_row
        for celda in ws[fila_encabezado]:
            celda.font = Font(bold=True, color="FFFFFF")
            celda.fill = PatternFill(start_color="007832", end_color="007832", fill_type="solid")

        detalles_venta = detalles_por_venta.get(venta["id_venta"], [])
        if not detalles_venta:
            ws.append(["Sin productos registrados", "", "", "", "", ""])
        else:
            for d in detalles_venta:
                precio = d.get("precio_venta")
                cantidad = d.get("cantidad")
                total_linea = float(precio) * float(cantidad) if precio is not None and cantidad is not None else 0
                ws.append([
                    d.get("nombre_producto") or "-",
                    cantidad,
                    d.get("simbolo") or "-",
                    float(precio) if precio is not None else "-",
                    total_linea,
                    d.get("estado_venta") or "-",
                ])

        ws.append([])

    fila_total = ws.max_row + 1
    ws.cell(row=fila_total, column=4, value="Total general:").font = Font(bold=True)
    ws.cell(row=fila_total, column=5, value=total_general).font = Font(bold=True)

    for columna in ws.columns:
        max_len = max((len(str(c.value)) if c.value else 0) for c in columna)
        ws.column_dimensions[columna[0].column_letter].width = max_len + 2

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

def generar_pdf_reporte_ventas(ventas: list, detalles: list) -> io.BytesIO:
    detalles_por_venta = _agrupar_detalles_por_venta(detalles)

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(letter),
        leftMargin=1.5 * cm,
        rightMargin=1.5 * cm,
        topMargin=1.5 * cm,
        bottomMargin=1.5 * cm,
    )
    styles = getSampleStyleSheet()
    elementos = [Paragraph("Informe de Ventas", styles["Title"]), Spacer(1, 12)]

    total_general = 0.0

    for venta in ventas:
        total_venta = float(venta.get("total_venta") or 0)
        total_general += total_venta

        resumen = [[
            f"Venta #{venta['id_venta']}",
            venta.get("nombre_comprador") or "-",
            str(venta.get("fecha_venta")) if venta.get("fecha_venta") else "-",
            venta.get("nombre_user") or "-",
            f"Total: ${total_venta:,.0f}",
        ]]
        tabla_resumen = Table(resumen, colWidths=[3 * cm, 5 * cm, 3 * cm, 3 * cm, 4 * cm])
        tabla_resumen.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#007832")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
        ]))
        elementos.append(tabla_resumen)

        detalles_venta = detalles_por_venta.get(venta["id_venta"], [])
        filas_detalle = [["Producto", "Cantidad", "Unidad", "Precio unitario", "Precio total", "Estado"]]
        if not detalles_venta:
            filas_detalle.append(["Sin productos registrados", "", "", "", "", ""])
        else:
            for d in detalles_venta:
                precio = d.get("precio_venta")
                cantidad = d.get("cantidad")
                total_linea = float(precio) * float(cantidad) if precio is not None and cantidad is not None else 0
                filas_detalle.append([
                    d.get("nombre_producto") or "-",
                    str(cantidad) if cantidad is not None else "-",
                    d.get("simbolo") or "-",
                    f"${float(precio):,.0f}" if precio is not None else "-",
                    f"${total_linea:,.0f}",
                    d.get("estado_venta") or "-",
                ])

        tabla_detalle = Table(filas_detalle, repeatRows=1, colWidths=[5 * cm, 2.5 * cm, 2 * cm, 3 * cm, 3 * cm, 3 * cm])
        tabla_detalle.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f3f4f6")),
            ("FONTSIZE", (0, 0), (-1, -1), 7),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ]))
        elementos.append(tabla_detalle)
        elementos.append(Spacer(1, 14))

    resumen_general = [
        ["Total de ventas", str(len(ventas))],
        ["Total general vendido", f"${total_general:,.0f}"],
    ]
    tabla_resumen_general = Table(resumen_general, colWidths=[6 * cm, 6 * cm])
    tabla_resumen_general.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, -1), colors.whitesmoke),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
    ]))
    elementos.append(tabla_resumen_general)

    doc.build(elementos)
    buffer.seek(0)
    return buffer

def generar_excel_reporte_tratamientos(tratamientos: list) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Tratamientos"

    ws.append(["Informe de Tratamientos"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([f"Total de registros: {len(tratamientos)}"])
    ws.append([])

    headers = [
        "Nombre lote", "Fecha inicio", "Fecha fin", "Cantidad",
        "Nombre producto", "Unidad", "Usuario", "Observación",
    ]
    ws.append(headers)
    fila_encabezado = ws.max_row
    for celda in ws[fila_encabezado]:
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = PatternFill(start_color="007832", end_color="007832", fill_type="solid")

    for t in tratamientos:
        ws.append([
            t.get("nombre_lote") or "-",
            str(t.get("fecha_inicio") or "-"),
            str(t.get("fecha_fin") or "-"),
            t.get("cantidad"),
            t.get("nombre_producto") or "-",
            t.get("simbolo") or "-",
            t.get("nombre_user") or "Sistema",
            t.get("observacion") or "",
        ])

    for columna in ws.columns:
        max_len = max((len(str(c.value)) if c.value else 0) for c in columna)
        ws.column_dimensions[columna[0].column_letter].width = max_len + 2

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

def generar_pdf_reporte_tratamientos(tratamientos: list) -> io.BytesIO:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(letter),
        leftMargin=1.5 * cm,
        rightMargin=1.5 * cm,
        topMargin=1.5 * cm,
        bottomMargin=1.5 * cm,
    )
    styles = getSampleStyleSheet()

    elementos = [Paragraph("Informe de Tratamientos", styles["Title"]), Spacer(1, 12)]

    filas = [["Lote", "Producto", "Fecha inicio", "Fecha fin", "Cantidad", "Usuario", "Observación"]]
    for t in tratamientos:
        cantidad = t.get("cantidad")
        simbolo = t.get("simbolo") or ""
        cant_convertida = t.get("cant_convertida")

        filas.append([
            t.get("nombre_lote") or "-",
            t.get("nombre_producto") or "-",
            str(t.get("fecha_inicio") or "-"),
            str(t.get("fecha_fin") or "-"),
            f"{cantidad} {simbolo}".strip(),
            t.get("nombre_user") or "Sistema",
            Paragraph(t.get("observacion") or "-", estilo_observaciones),
        ])

    tabla = Table(
        filas,
        repeatRows=1,
        colWidths=[2.5 * cm, 4 * cm, 3 * cm, 3 * cm, 2.5 * cm, 2.5 * cm, 2.5 * cm, 5 * cm],
    )
    tabla.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f3f4f6")),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]))
    elementos.append(tabla)
    elementos.append(Spacer(1, 18))
    elementos.append(Paragraph(f"Total de registros: {len(tratamientos)}", styles["Normal"]))

    doc.build(elementos)
    buffer.seek(0)
    return buffer

def generar_excel_reporte_ventas_platos(venta_platos: list) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Ventas de platos"

    ws.append(["Informe de Ventas de Platos"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([f"Total de registros: {len(venta_platos)}"])
    ws.append([])

    headers = [
        "Nombre plato", "Cantidad", "Precio", "Fecha de venta",
    ]
    ws.append(headers)
    fila_encabezado = ws.max_row
    for celda in ws[fila_encabezado]:
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = PatternFill(start_color="007832", end_color="007832", fill_type="solid")

    for t in venta_platos:
        ws.append([
            t.get("nombre_plato") or "-",
            t.get("cantidad"),
            t.get("precio") or "-",
            str(t.get("fecha_venta") or "-"),
        ])

    for columna in ws.columns:
        max_len = max((len(str(c.value)) if c.value else 0) for c in columna)
        ws.column_dimensions[columna[0].column_letter].width = max_len + 2

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

def generar_pdf_reporte_ventas_platos(venta_platos: list) -> io.BytesIO:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(letter),
        leftMargin=1.5 * cm,
        rightMargin=1.5 * cm,
        topMargin=1.5 * cm,
        bottomMargin=1.5 * cm,
    )
    styles = getSampleStyleSheet()

    elementos = [Paragraph("Informe de Ventas de Platos", styles["Title"]), Spacer(1, 12)]

    filas = [["Nombre plato", "Cantidad", "Precio", "Fecha de venta",]]
    for t in venta_platos:

        filas.append([
            t.get("nombre_plato") or "-",
            t.get("cantidad"),
            t.get("precio") or "-",
            str(t.get("fecha_venta") or "-"),
        ])

    tabla = Table(
        filas,
        repeatRows=1,
        colWidths=[2.5 * cm, 2 * cm, 3 * cm, 3 * cm],
    )
    tabla.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f3f4f6")),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]))
    elementos.append(tabla)
    elementos.append(Spacer(1, 18))
    elementos.append(Paragraph(f"Total de registros: {len(venta_platos)}", styles["Normal"]))

    doc.build(elementos)
    buffer.seek(0)
    return buffer

def generar_excel_reporte_maquina(reporte: dict) -> io.BytesIO:
    encabezado, movimientos = _normalizar_reporte_maquina(reporte)

    wb = Workbook()
    ws = wb.active
    ws.title = "Encabezado"

    ws.append(["Informe de maquina", encabezado["nombre_maq"]])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])
    ws.append(["Marca", str(encabezado["marca"])])
    ws.append(["Modelo", str(encabezado["modelo"])])
    ws.append(["Tipo maquina", str(encabezado["tipo_maq"])])
    ws.append(["Fecha de compra", str(encabezado["fecha_compra"])])

    for fila in ws.iter_rows(min_row=3, max_row=9, min_col=1, max_col=1):
        for celda in fila:
            celda.font = Font(bold=True)

    ws_mov = wb.create_sheet("Historial de estados")
    ws_mov.append(["Estado", "Fecha de registro", "Responsable", "Observaciones"])
    for celda in ws_mov[1]:
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = PatternFill(start_color="007832", end_color="007832", fill_type="solid")

    for m in movimientos:
        ws_mov.append([
            m.get("estado_actual", m.get("estado", "")),
            str(m.get("fecha_cambio", m.get("fecha_registro", ""))),
            m.get("nombre_user", m.get("responsable", "")),
            m.get("observaciones"),
        ])

    for columna in ws_mov.columns:
        max_len = max((len(str(c.value)) if c.value else 0) for c in columna)
        ws_mov.column_dimensions[columna[0].column_letter].width = max_len + 2

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

def generar_pdf_reporte_maquina(reporte: dict) -> io.BytesIO:
    encabezado, movimientos = _normalizar_reporte_maquina(reporte)

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    styles = getSampleStyleSheet()

    # Estilo centrado para el título de historial de estados
    estilo_titulo_centrado = ParagraphStyle(
        "TituloCentrado",
        parent=styles["Heading2"],
        alignment=TA_CENTER,
    )

    elementos = []

    elementos.append(Paragraph(f"Informe de maquina: {encabezado['nombre_maq']}", styles["Title"]))
    elementos.append(Spacer(1, 12))

    datos_encabezado = [
        ["Nombre de la maquina", str(encabezado["nombre_maq"])],
        ["Marca", str(encabezado["marca"])],
        ["Modelo", str(encabezado["modelo"])],
        ["Tipo maquina", str(encabezado["tipo_maq"])],
        ["Fecha de compra", str(encabezado["fecha_compra"])],
    ]
    tabla_encabezado = Table(datos_encabezado, colWidths=[6 * cm, 6 * cm])
    tabla_encabezado.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, -1), colors.whitesmoke),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
    ]))
    elementos.append(tabla_encabezado)
    elementos.append(Spacer(1, 20))

    # Título "Historial de estados" centrado
    elementos.append(Paragraph("Historial de estados", estilo_titulo_centrado))
    elementos.append(Spacer(1, 6))

    filas = [["Estado", "Fecha de registro", "Responsable", "Observaciones"]]
    for m in movimientos:
        filas.append([
            m.get("estado_actual", m.get("estado", "")),
            str(m.get("fecha_cambio", m.get("fecha_registro", ""))),
            m.get("nombre_user", m.get("responsable", "")),
            Paragraph(m.get("observaciones") or "", estilo_observaciones),
        ])

    tabla_movs = Table(
        filas,
        repeatRows=1,
        colWidths=[3 * cm, 3.5 * cm, 3.5 * cm, 6.5 * cm],
    )
    tabla_movs.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f3f4f6")),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))
    elementos.append(tabla_movs)

    doc.build(elementos)
    buffer.seek(0)
    return buffer

def generar_excel_reporte_general_maquina(maquinaria: list) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Maquinaria"

    ws.append(["Informe de Maquinaria"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([f"Total de registros: {len(maquinaria)}"])
    ws.append([])

    headers = [
        "Nombre máquina", "Tipo", "Marca/Modelo", "N° Serie",
        "Fecha de compra", "Estado", "Ubicación", "Observación", "Fecha de retiro",
    ]
    ws.append(headers)
    fila_encabezado = ws.max_row
    for celda in ws[fila_encabezado]:
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = PatternFill(start_color="007832", end_color="007832", fill_type="solid")

    for t in maquinaria:
        ws.append([
            t.get("nombre_maq") or "-",
            t.get("tipo_maq") or "-",
            t.get("marca") + "/" + t.get("modelo") or "-",
            t.get("num_serie") or "-",
            str(t.get("fecha_compra") or "-"),
            t.get("estado") or "-",
            t.get("ubicacion") or "-",
            t.get("observaciones") or "-",
            str(t.get("fecha_de_baja") or "-"),
        ])

    for columna in ws.columns:
        max_len = max((len(str(c.value)) if c.value else 0) for c in columna)
        ws.column_dimensions[columna[0].column_letter].width = max_len + 2

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

def generar_pdf_reporte_general_maquina(maquinaria: list) -> io.BytesIO:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(letter),
        leftMargin=1.5 * cm,
        rightMargin=1.5 * cm,
        topMargin=1.5 * cm,
        bottomMargin=1.5 * cm,
    )
    styles = getSampleStyleSheet()
    estilo_observaciones = styles["BodyText"].clone("Observaciones maquinaria")
    estilo_observaciones.fontSize = 6
    estilo_observaciones.leading = 7

    elementos = [Paragraph("Informe de maquinaria", styles["Title"]), Spacer(1, 12)]

    filas = [["Nombre máquina", "Tipo", "Marca/Modelo", "N° Serie", "Fecha de compra", "Estado", "Ubicación", "Observación", "Fecha de retiro"]]
    for t in maquinaria:

        filas.append([
            t.get("nombre_maq") or "-",
            t.get("tipo_maq") or "-",
            t.get("marca") + "/" + t.get("modelo") or "-",
            t.get("num_serie") or "-",
            str(t.get("fecha_compra") or "-"),
            t.get("estado") or "-",
            t.get("ubicacion") or "-",
            Paragraph(t.get("observaciones") or "-", estilo_observaciones),
            str(t.get("fecha_de_baja") or "-"),
        ])

    tabla = Table(
        filas,
        repeatRows=1,
        colWidths=[3.5 * cm, 2.5 * cm, 3 * cm, 2.5 * cm, 2.5 * cm, 2.5 * cm, 2.5 * cm, 3 * cm,],
    )
    tabla.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f3f4f6")),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]))
    elementos.append(tabla)
    elementos.append(Spacer(1, 18))
    elementos.append(Paragraph(f"Total de registros: {len(maquinaria)}", styles["Normal"]))

    doc.build(elementos)
    buffer.seek(0)
    return buffer

def generar_excel_reporte_soli_insumo(solicitud: list) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Solicitudes insumos"

    ws.append(["Informe de Solicitudes de Insumos"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([f"Total de registros: {len(solicitud)}"])
    ws.append([])

    headers = [
        "Solicitante", "Ficha", "Fecha solicitud", "Fecha entrega", "Fecha devolución", 
        "Cant. solicitada", "Cant. a devolver", "Producto", 
        "Categoria", "Estado",
    ]
    ws.append(headers)
    fila_encabezado = ws.max_row
    for celda in ws[fila_encabezado]:
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = PatternFill(start_color="007832", end_color="007832", fill_type="solid")

    for t in solicitud:
        ws.append([
            t.get("solicitante") or "-",
            t.get("ficha") or "-",
            str(t.get("fecha_solicitud")or "-"),
            str(t.get("fecha_entrega")or "-"),
            str(t.get("fecha_devolucion")or "-"),
            str(t.get("cantidad_in") or "-"),
            str(t.get("cant_devolver") or "-"),
            t.get("nombre_producto") or "-",
            t.get("nombre_tipo") or "-",
            t.get("estado_solicitud") or "-",
        ])

    for columna in ws.columns:
        max_len = max((len(str(c.value)) if c.value else 0) for c in columna)
        ws.column_dimensions[columna[0].column_letter].width = max_len + 2

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

def generar_pdf_reporte_soli_insumo(solicitud: list) -> io.BytesIO:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(letter),
        leftMargin=1.5 * cm,
        rightMargin=1.5 * cm,
        topMargin=1.5 * cm,
        bottomMargin=1.5 * cm,
    )
    styles = getSampleStyleSheet()
    estilo_observaciones.fontSize = 6
    estilo_observaciones.leading = 7

    elementos = [Paragraph("Informe de solicitudes de insumos", styles["Title"]), Spacer(1, 12)]

    filas = [["Solicitante", "Ficha", "Fecha solicitud", "Fecha entrega", "Fecha devolución", 
            "Cant. solicitada", "Cant. a devolver", "Producto", 
            "Categoria", "Estado"]]
    
    for t in solicitud:
        filas.append([
            t.get("solicitante") or "-",
            t.get("ficha") or "-",
            str(t.get("fecha_solicitud")or "-"),
            str(t.get("fecha_entrega")or "-"),
            str(t.get("fecha_devolucion")or "-"),
            str(t.get("cantidad_in") or "-"),
            str(t.get("cant_devolver") or "-"),
            t.get("nombre_producto") or "-",
            t.get("nombre_tipo") or "-",
            t.get("estado_solicitud") or "-",
        ])

    tabla = Table(
        filas,
        repeatRows=1,
        colWidths=[3.5 * cm, 2.5 * cm, 2.5 * cm, 2.5 * cm, 2.5 * cm, 2 * cm, 2 * cm, 2.5 * cm, 2.5 * cm, 2.5 * cm],
    )
    tabla.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f3f4f6")),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]))
    elementos.append(tabla)
    elementos.append(Spacer(1, 18))
    elementos.append(Paragraph(f"Total de registros: {len(solicitud)}", styles["Normal"]))

    doc.build(elementos)
    buffer.seek(0)
    return buffer

def generar_excel_reporte_comercializacion(comercializacion: list) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Comercialización de productos"

    ws.append(["Informe de Comercializaciones"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([f"Total de registros: {len(comercializacion)}"])
    ws.append([])

    headers = [
        "Producto", "Lote", "Fecha", "Cantidad", "Lugar", "Estado de vendido", 
        "Cant. no vendida",
    ]
    ws.append(headers)
    fila_encabezado = ws.max_row
    for celda in ws[fila_encabezado]:
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = PatternFill(start_color="007832", end_color="007832", fill_type="solid")

    for t in comercializacion:
        ws.append([
            t.get("nombre_producto") or "-",
            t.get("sublote") or "-",
            t.get("fecha_comercializacion") or "-",
            str(t.get("cantidad") or "-") + "/" + str(t.get("simbolo") or "-"),
            t.get("lugar_comercializacion") or "-",
            "Vendió todo" if t.get("vendio_todo") else "No vendió todo",
            str(t.get("cant_no_vendida") or "-"),
        ])

    for columna in ws.columns:
        max_len = max((len(str(c.value)) if c.value else 0) for c in columna)
        ws.column_dimensions[columna[0].column_letter].width = max_len + 2

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

def generar_pdf_reporte_comercializacion(comercializacion: list) -> io.BytesIO:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(letter),
        leftMargin=1.5 * cm,
        rightMargin=1.5 * cm,
        topMargin=1.5 * cm,
        bottomMargin=1.5 * cm,
    )
    styles = getSampleStyleSheet()
    estilo_observaciones.fontSize = 6
    estilo_observaciones.leading = 7

    elementos = [Paragraph("Informe de comercializaciones", styles["Title"]), Spacer(1, 12)]

    filas = [["Producto", "Lote", "Fecha", "Cantidad", "Lugar", "Estado de vendido", "Cant. no vendida"]]
    
    for t in comercializacion:
        filas.append([
            t.get("nombre_producto") or "-",
            t.get("sublote") or "-",
            t.get("fecha_comercializacion") or "-",
            str(t.get("cantidad") or "-") + "/" + str(t.get("simbolo") or "-"),
            t.get("lugar_comercializacion") or "-",
            "Vendió todo" if t.get("vendio_todo") else "No vendió todo",
            str(t.get("cant_no_vendida") or "-"),
        ])

    tabla = Table(
        filas,
        repeatRows=1,
        colWidths=[3.5 * cm, 2 * cm, 2.5 * cm, 2 * cm, 4 * cm, 2.5 * cm, 2.5 * cm],
    )
    tabla.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f3f4f6")),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]))
    elementos.append(tabla)
    elementos.append(Spacer(1, 18))
    elementos.append(Paragraph(f"Total de registros: {len(comercializacion)}", styles["Normal"]))

    doc.build(elementos)
    buffer.seek(0)
    return buffer