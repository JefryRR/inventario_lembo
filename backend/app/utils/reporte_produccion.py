import io
from openpyxl import Workbook   # type: ignore
from openpyxl.styles import Font, PatternFill  # type: ignore
from reportlab.lib import colors  # type: ignore
from reportlab.lib.pagesizes import letter, landscape  # type: ignore
from reportlab.lib.units import cm # type: ignore
from reportlab.lib.styles import getSampleStyleSheet  # type: ignore
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer # type: ignore
from reportlab.platypus import Paragraph  # type: ignore
from typing import Union
from datetime import datetime, date

styles = getSampleStyleSheet()

estilo_observaciones = styles["BodyText"].clone("Observaciones")
estilo_observaciones.fontSize = 6
estilo_observaciones.leading = 7

def _formatear_fecha_hora(valor) -> Paragraph:
    if valor in (None, "-", ""):
        return Paragraph("-", estilo_observaciones)

    fecha_dt = None
    if isinstance(valor, (datetime, date)):
        fecha_dt = valor
    elif isinstance(valor, str):
        texto = valor.replace("T", " ")
        for fmt in ("%Y-%m-%d %H:%M:%S.%f", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
            try:
                fecha_dt = datetime.strptime(texto, fmt)
                break
            except ValueError:
                continue

    if fecha_dt is None:
        return Paragraph(str(valor), estilo_observaciones)

    fecha_str = fecha_dt.strftime("%d/%m/%Y")
    # Si es un date puro (sin hora) o la hora es 00:00:00, no mostramos hora
    if isinstance(fecha_dt, datetime) and not (fecha_dt.hour == 0 and fecha_dt.minute == 0 and fecha_dt.second == 0):
        hora_str = fecha_dt.strftime("%H:%M")
        return Paragraph(f"{fecha_str}<br/>{hora_str}", estilo_observaciones)

    return Paragraph(fecha_str, estilo_observaciones)

def generar_pdf_rep_gral_produccion(reporte: list) -> io.BytesIO:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(letter),
        leftMargin=1.5 * cm,
        rightMargin=1.5 * cm,
        topMargin=1.5 * cm,
        bottomMargin=1.5 * cm,
    )
    styles = getSampleStyleSheet()
    elementos = []

    elementos.append(Paragraph("Informe de Producción", styles["Title"]))
    elementos.append(Spacer(1, 12))

    total_produccion = 0.0
    filas: list[list[Union[str, Paragraph]]] = [["Producto", "cantidad", "F. registro", "F. vencimiento", "Categoría", "Especie", "Lote", "Costo unit.", "Costo total.", "Estado"]]
    for p in reporte:
        valor_unitario = p.get("valor_unitario")
        valor_total = _calcular_total_prod(p)
        total_produccion += valor_total
        filas.append([
            Paragraph(p.get("nombre_producto") or "", estilo_observaciones),
            f'{p.get("cantidad") or "0"} {p.get("simbolo") or ""}'.strip(),
            _formatear_fecha_hora(p.get("fecha_ingreso")),
            _formatear_fecha_hora(p.get("fecha_vencimiento")),
            p.get("nombre_categoria") or "",
            p.get("nombre_especie") or "",
            p.get("nombre_lote") or "",
            f"${float(valor_unitario):,.0f}" if valor_unitario not in (None, "-") else "-",
            f"${valor_total:,.0f}",
            Paragraph(p.get("nivel_alerta") or "N/A", estilo_observaciones)
        ])

    tabla = Table(
        filas,
        repeatRows=1,
        colWidths=[
            2.5 * cm,
            2 * cm,
            2 * cm,
            2 * cm,
            2.3 * cm,
            2 * cm,
            2 * cm,
            2.3 * cm,
            2.3 * cm,
            4 * cm,
        ]
    )
    tabla.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f3f4f6")),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]))
    elementos.append(tabla)
    elementos.append(Spacer(1, 18))

    resumen = [
        ["Total de registros", str(len(reporte))],
        ["Valor total de producción", f"${total_produccion:,.0f}"],
    ]
    tabla_resumen = Table(resumen, colWidths=[6 * cm, 6 * cm])
    tabla_resumen.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, -1), colors.whitesmoke),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
    ]))
    elementos.append(tabla_resumen)

    doc.build(elementos)
    buffer.seek(0)
    return buffer

def generar_excel_rep_gral_produccion(produccion: list) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Producción"

    ws.append(["Informe de Producción"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([f"Total de registros: {len(produccion)}"])
    ws.append([])

    headers = [
        "Producto", "cantidad", "F. registro", "F. vencimiento", "Categoría", "Especie", "Lote", 
        "Costo unit.", "Costo total.", "Estado"
    ]
    ws.append(headers)
    fila_encabezado = ws.max_row
    for celda in ws[fila_encabezado]:
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = PatternFill(start_color="007832", end_color="007832", fill_type="solid")
    total_produccion = 0.0
    for p in produccion:
        valor_total = _calcular_total_prod(p)
        total_produccion += valor_total
    
        ws.append([
            p.get("nombre_producto"),
            f'{p.get("cantidad") or "0"} {p.get("simbolo") or ""}'.strip(),
            p.get("fecha_ingreso"),
            p.get("fecha_vencimiento"),
            p.get("nombre_categoria") or "-",
            p.get("nombre_especie") or "-",
            p.get("nombre_lote") or "-",
            float(p.get("valor_unitario")) if p.get("valor_unitario") not in (None, "-") else "-",
            valor_total,
            p.get("nivel_alerta") or "N/A"
        ])

    ws.append([])
    fila_total = ws.max_row + 1
    ws.cell(row=fila_total, column=8, value="Valor total producción:").font = Font(bold=True)
    ws.cell(row=fila_total, column=9, value=total_produccion).font = Font(bold=True)

    for columna in ws.columns:
        max_len = max((len(str(c.value)) if c.value else 0) for c in columna)
        ws.column_dimensions[columna[0].column_letter].width = max_len + 2

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

def _calcular_total_prod(produccion: dict) -> float:
    valor_unitario = produccion.get("valor_unitario")
    cantidad = produccion.get("cantidad")
    if valor_unitario in (None, "-") or cantidad in (None, "-"):
        return 0.0
    return float(valor_unitario) * float(cantidad)

def generar_excel_reporte_produccion(reporte: dict) -> io.BytesIO:
    encabezado = reporte["encabezado"]
    movimientos = reporte["movimientos"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Encabezado"

    ws.append(["Informe de Producción", encabezado["nombre_producto"]])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])
    ws.append(["Fecha de ingreso", str(encabezado["fecha_ingreso"])])
    ws.append(["Fecha de vencimiento", str(encabezado["fecha_vencimiento"])])
    ws.append(["Costo unitario", float(encabezado["valor_unitario"])])
    ws.append(["Nombre lote", str(encabezado["nombre_lote"])])
    ws.append(["Cantidad inicial", encabezado["cantidad_inicial"]])
    ws.append(["Unidad de medida", str(encabezado["simbolo"])])
    ws.append(["Stock actual", encabezado["stock_actual"]])
    ws.append(["Total Vendido", encabezado["total_vendido"]])
    ws.append(["Total perdido", encabezado["total_perdido"]])

    for fila in ws.iter_rows(min_row=3, max_row=9, min_col=1, max_col=1):
        for celda in fila:
            celda.font = Font(bold=True)

    ws_mov = wb.create_sheet("Movimientos")
    ws_mov.append(["Tipo", "Observaciones", "Cantidad", "Unidad", "Valor", "Motivo", "Fecha"])
    for celda in ws_mov[1]:
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = PatternFill(start_color="007832", end_color="007832", fill_type="solid")

    for m in movimientos:
        ws_mov.append([
            m.get("tipo"),
            m.get("referencia"),
            m.get("cantidad"),
            encabezado.get("simbolo"),
            float(m["valor"]) if m.get("valor") not in (None, "-") else "-",
            m.get("motivo"),
            str(m.get("fecha")),
        ])

    for columna in ws_mov.columns:
        max_len = max((len(str(c.value)) if c.value else 0) for c in columna)
        ws_mov.column_dimensions[columna[0].column_letter].width = max_len + 2

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

def generar_pdf_reporte_produccion(reporte: dict) -> io.BytesIO:
    encabezado = reporte["encabezado"]
    movimientos = reporte["movimientos"]

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    styles = getSampleStyleSheet()
    elementos = []

    elementos.append(Paragraph(f"Informe de Producción: {encabezado['nombre_producto']}", styles["Title"]))
    elementos.append(Spacer(1, 12))

    datos_encabezado = [
        ["Nombre del producto", str(encabezado["nombre_producto"])],
        ["Fecha de ingreso", str(encabezado["fecha_ingreso"])],
        ["Fecha de vencimiento", str(encabezado["fecha_vencimiento"])],
        ["Costo unitario", f"${float(encabezado['valor_unitario']):,.0f}"],
        ["Nombre lote", str(encabezado["nombre_lote"])],
        ["Cantidad inicial", f"{encabezado['cantidad_inicial']} {encabezado['simbolo']}"],
        ["Stock actual", f"{encabezado['stock_actual']} {encabezado['simbolo']}"],
        ["Total Vendido", f"{encabezado['total_vendido']} {encabezado['simbolo']}"],
        ["Total perdido", f"{encabezado['total_perdido']} {encabezado['simbolo']}"],
    ]
    tabla_encabezado = Table(datos_encabezado, colWidths=[6 * cm, 6 * cm])
    tabla_encabezado.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, -1), colors.whitesmoke),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
    ]))
    elementos.append(tabla_encabezado)
    elementos.append(Spacer(1, 20))

    elementos.append(Paragraph("Movimientos", styles["Heading2"]))
    filas = [["Tipo", "Observaciones", "Cantidad", "Valor", "Motivo", "Fecha"]]
    for m in movimientos:
        valor = m.get("valor")
        filas.append([
            m.get("tipo", ""),
            m.get("referencia") or "",
            str(m.get("cantidad", "")),
            f"${float(valor):,.0f}" if valor not in (None, "-") else "-",
            m.get("motivo") or "",
            str(m.get("fecha", "")),
        ])

    tabla_movs = Table(filas, repeatRows=1)
    tabla_movs.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f3f4f6")),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))
    elementos.append(tabla_movs)

    doc.build(elementos)
    buffer.seek(0)
    return buffer
